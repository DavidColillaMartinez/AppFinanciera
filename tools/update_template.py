import openpyxl
import sys

XLSX = r'C:\Users\david\Documents\TEST\AppFinanzas\plantilla_base_finanzas_app.xlsx'

wb = openpyxl.load_workbook(XLSX)

# 1) Cuentas: anadir columna "rol" entre "tipo" y "moneda"
ws = wb['Cuentas']
old_headers = [c.value for c in ws[1]]
print('Cuentas headers actuales:', old_headers)
# Capturar todas las filas existentes
rows = []
for r in range(1, ws.max_row + 1):
    rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

# Nuevo header: insertar 'rol' en posicion 4 (entre 'tipo' y 'moneda')
new_header = ['cuentaId', 'nombre', 'tipo', 'rol', 'moneda', 'saldoInicial',
              'saldoActualManual', 'incluirDashboard', 'activo', 'color',
              'notas', 'createdAt', 'updatedAt']
# Mapeo de col antigua -> col nueva
old_idx_to_new = {1: 1, 2: 2, 3: 3, 4: 5, 5: 6, 6: 7, 7: 8, 8: 9, 9: 10,
                  10: 11, 11: 12, 12: 13}
new_rows = [new_header]
for r_idx, old_row in enumerate(rows[1:], start=2):  # saltar header
    new_row = [None] * 13
    for old_c, new_c in old_idx_to_new.items():
        if old_c - 1 < len(old_row):
            new_row[new_c - 1] = old_row[old_c - 1]
    # Asignar rol por defecto segun el nombre
    nombre = str(new_row[1] or '').lower()
    if 'principal' in nombre:
        new_row[3] = 'general'
    elif 'ahorro' in nombre:
        new_row[3] = 'ahorro'
    else:
        new_row[3] = 'diario'
    new_rows.append(new_row)

# Anadir una cuenta de ahorro de ejemplo si no existe
existing_ids = [r[0] for r in rows[1:]]
if not any(str(i) == 'cta-ahorro' for i in existing_ids):
    new_rows.append([
        'cta-ahorro', 'Cuenta de ahorro', 'Banco', 'ahorro', 'EUR', 0, 0,
        'No', 'Si', '#10B981',
        'Cuenta donde se reservan los ahorros. Aparece en el libro de ahorro.',
        '', '',
    ])

# Limpiar hoja y reescribir
for r in range(ws.max_row, 0, -1):
    ws.delete_rows(r, 1)
for r_idx, row in enumerate(new_rows, start=1):
    for c_idx, val in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=val)

# 2) Mov_reservas: anadir mesClave, tipoDestino, destinoId, updatedAt
ws = wb['Mov_reservas']
old_headers = [c.value for c in ws[1]]
print('Mov_reservas headers actuales:', old_headers)

new_mov_header = ['id', 'fecha', 'mesClave', 'tipoDestino', 'destinoId',
                  'reservaId', 'tipoMovimiento', 'importe', 'cuentaOrigen',
                  'cuentaDestino', 'notas', 'createdAt', 'updatedAt']
# Limpiar hoja
for r in range(ws.max_row, 0, -1):
    ws.delete_rows(r, 1)
# Header
for c_idx, h in enumerate(new_mov_header, start=1):
    ws.cell(row=1, column=c_idx, value=h)
# Fila de ejemplo
ws.cell(row=2, column=1, value='LEDGER-EXAMPLE-001')
ws.cell(row=2, column=2, value='2026-01-15')
ws.cell(row=2, column=3, value='2026-01')
ws.cell(row=2, column=4, value='reserva')
ws.cell(row=2, column=5, value='res-imprevistos')
ws.cell(row=2, column=6, value='res-imprevistos')
ws.cell(row=2, column=7, value='aporte')
ws.cell(row=2, column=8, value=75)
ws.cell(row=2, column=9, value='cta-principal')
ws.cell(row=2, column=10, value='cta-ahorro')
ws.cell(row=2, column=11, value='Ejemplo: primer aporte del ano a la reserva de imprevistos.')
ws.cell(row=2, column=12, value='2026-01-15T00:00:00.000Z')
ws.cell(row=2, column=13, value='2026-01-15T00:00:00.000Z')

# 3) Config: reescribir con las claves actualizadas
ws = wb['Config']
new_config = [
    ('Clave', 'Valor', 'Descripcion'),
    ('templateVersion', '1.1.0', 'Version de estructura de la plantilla. Se actualiza ante cambios incompatibles.'),
    ('appMinVersion', '1.0.0', 'Version minima de la app que entiende esta plantilla.'),
    ('currency', 'EUR', 'Moneda principal.'),
    ('year', '2026', 'Ano activo para dashboard/calendario.'),
    ('month', '1', 'Mes activo numerico: 1-12.'),
    ('firstDayOfWeek', '1', 'Primer dia de la semana: 1=lunes, 7=domingo.'),
    ('locale', 'es-ES', 'Locale para formateo de numeros y fechas.'),
    ('salary.enabled', 'false', 'Activa la generacion automatica del movimiento de nomina.'),
    ('salary.type', 'fixed', 'Tipo: fixed | variable.'),
    ('salary.fixedAmount', '0', 'Importe de la nomina cuando type=fixed.'),
    ('salary.day', '1', 'Dia del mes en que se anade la nomina.'),
    ('salary.destinationAccount', '', 'Nombre de la cuenta destino del movimiento de nomina.'),
    ('salary.description', 'Nomina', 'Concepto por defecto del movimiento de nomina.'),
    ('salary.updatedAt', '', 'Marca de tiempo de la ultima actualizacion de la nomina.'),
    ('fixed.confirmed.YYYY-MM', '', 'Patron: clave de confirmacion de gastos fijos por mes. La app la escribe automaticamente. No editar a mano.'),
    ('schemaLocked', 'Si', 'No cambiar nombres de columnas si la app esta conectada.'),
    ('ownerName', '', 'Opcional: nombre del usuario.'),
    ('appMode', 'Sheet as database', 'La Google Sheet actua como base de datos ligera.'),
]
for r in range(ws.max_row, 0, -1):
    ws.delete_rows(r, 1)
for r_idx, row in enumerate(new_config, start=1):
    for c_idx, val in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=val)

# 4) Actualizar 00_LEEME
ws = wb['00_LEEME']
new_leeme = [
    ['PLANTILLA BASE - FINANZAS PERSONALES', 'Plantilla compatible con la app (templateVersion 1.1.0).'],
    ['Objetivo', 'Base generica para que cada usuario la suba a Google Sheets y conecte la app.'],
    ['Como usarla', '1) Subir este XLSX a Google Drive. 2) Abrir con Google Sheets. 3) Conectar la app desde el onboarding.'],
    ['Hojas oficiales', 'Config, Movimientos, Categorias, Cuentas, Gastos_fijos, Pagos_futuros, Objetivos, Reservas, Mov_reservas, Pagos_aplazados.'],
    ['Hojas heredadas', 'Dashboard, Vista_mes, AppsScript: compatibles con el Apps Script legado (no usado por la app oficial).'],
    ['Cuentas.rol', 'Valores permitidos: diario, fijos, ahorro, general. La app usa el rol para calcular Disponible.'],
    ['Mov_reservas', 'Libro de ahorro: tipoDestino=reserva|objetivo|pago_futuro. tipoMovimiento=aporte|retirada.'],
    ['Mov_reservas.Eliminado', 'El sentinel "Eliminado" en tipoMovimiento se ignora en los calculos.'],
    ['Importante', 'No renombres hojas ni columnas. Para limpiar registros, usa el campo deletedAt en Movimientos.'],
]
for r in range(ws.max_row, 0, -1):
    ws.delete_rows(r, 1)
for r_idx, row in enumerate(new_leeme, start=1):
    for c_idx, val in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=val)

wb.save(XLSX)
print('OK: template actualizado a 1.1.0')
